# --- Attendance Routes ---
# GET endpoints for attendance records, today's list, stats, and analytics.
# PATCH endpoint for toggling attendance status.
# GET endpoints for Excel exports (daily attendance, student summary).

import io
from flask import Blueprint, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models.attendance import (
    get_attendance_records, get_today_attendance,
    get_attendance_stats, get_analytics_data,
    update_attendance_status, get_attendance_for_export,
    get_student_attendance_summary
)
from routes.auth import auth_required

attendance_bp = Blueprint('attendance', __name__)


# --- GET /api/attendance ---
@attendance_bp.route('/api/attendance', methods=['GET'])
@auth_required
def list_attendance():
    records = get_attendance_records(
        filter_date=request.args.get('date'),
        student_id=request.args.get('student_id'),
        section=request.args.get('section'),
        branch=request.args.get('branch'),
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to')
    )
    return jsonify(records), 200


# --- GET /api/attendance/today ---
@attendance_bp.route('/api/attendance/today', methods=['GET'])
@auth_required
def today_attendance():
    records = get_today_attendance()
    return jsonify(records), 200


# --- GET /api/attendance/stats ---
@attendance_bp.route('/api/attendance/stats', methods=['GET'])
@auth_required
def attendance_stats():
    stats = get_attendance_stats()
    return jsonify(stats), 200


# --- GET /api/attendance/analytics ---
@attendance_bp.route('/api/attendance/analytics', methods=['GET'])
@auth_required
def attendance_analytics():
    data = get_analytics_data(
        date_from=request.args.get('date_from'),
        date_to=request.args.get('date_to'),
        section=request.args.get('section'),
        branch=request.args.get('branch')
    )
    return jsonify(data), 200


# --- PATCH /api/attendance/<id>/status ---
@attendance_bp.route('/api/attendance/<int:attendance_id>/status', methods=['PATCH'])
@auth_required
def toggle_attendance_status(attendance_id):
    """Toggle attendance status between Present and Absent."""
    data = request.get_json()
    if not data or 'status' not in data:
        return jsonify({"error": "Status is required"}), 400

    new_status = data['status']
    if new_status not in ('Present', 'Absent'):
        return jsonify({"error": "Status must be 'Present' or 'Absent'"}), 400

    updated = update_attendance_status(attendance_id, new_status)
    if not updated:
        return jsonify({"error": "Attendance record not found"}), 404

    return jsonify({"message": f"Status updated to {new_status}", "status": new_status}), 200


# --- Excel Styling Helpers ---
def _style_excel_header(ws, headers, row=1):
    """Apply premium dark-themed styling to header row."""
    header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='FF5722', end_color='FF5722', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border


def _style_excel_data(ws, start_row, end_row, num_cols):
    """Apply styling to data rows."""
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    present_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    absent_fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')

    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Color the status column if present
        status_cell = ws.cell(row=row_idx, column=num_cols)
        if status_cell.value == 'Present':
            status_cell.fill = present_fill
            status_cell.font = Font(name='Calibri', size=10, bold=True, color='2E7D32')
        elif status_cell.value == 'Absent':
            status_cell.fill = absent_fill
            status_cell.font = Font(name='Calibri', size=10, bold=True, color='C62828')


# --- GET /api/attendance/export/daily ---
@attendance_bp.route('/api/attendance/export/daily', methods=['GET'])
@auth_required
def export_daily_attendance():
    """Export daily attendance as Excel file."""
    filter_date = request.args.get('date')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    section = request.args.get('section')
    branch = request.args.get('branch')

    data = get_attendance_for_export(
        filter_date=filter_date,
        date_from=date_from,
        date_to=date_to,
        section=section,
        branch=branch
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Attendance"

    # --- Title Row ---
    title = f"Daily Attendance Report"
    if filter_date:
        title += f" — {filter_date}"
    elif date_from and date_to:
        title += f" — {date_from} to {date_to}"
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='FF5722')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Headers ---
    headers = ['Date', 'Student Name', 'USN (Roll No.)', 'Branch', 'Section', 'Time', 'Status']
    _style_excel_header(ws, headers, row=3)

    # --- Data ---
    for idx, record in enumerate(data, start=4):
        ws.cell(row=idx, column=1, value=record['date'])
        ws.cell(row=idx, column=2, value=record['name'])
        ws.cell(row=idx, column=3, value=record['roll_number'])
        ws.cell(row=idx, column=4, value=record['branch'])
        ws.cell(row=idx, column=5, value=record['section'])
        ws.cell(row=idx, column=6, value=record['time'])
        ws.cell(row=idx, column=7, value=record['status'])

    if data:
        _style_excel_data(ws, 4, 3 + len(data), 7)

    # --- Auto-fit column widths ---
    for col_idx in range(1, 8):
        max_len = 0
        for row_idx in range(3, 4 + len(data)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # --- Send file ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "daily_attendance"
    if filter_date:
        filename += f"_{filter_date}"
    elif date_from and date_to:
        filename += f"_{date_from}_to_{date_to}"
    filename += ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# --- GET /api/attendance/export/summary ---
@attendance_bp.route('/api/attendance/export/summary', methods=['GET'])
@auth_required
def export_attendance_summary():
    """Export student attendance summary as Excel file."""
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    section = request.args.get('section')
    branch = request.args.get('branch')
    min_pct = request.args.get('min_percentage')

    min_percentage = None
    if min_pct:
        try:
            min_percentage = float(min_pct)
        except ValueError:
            return jsonify({"error": "min_percentage must be a number"}), 400

    data = get_student_attendance_summary(
        date_from=date_from,
        date_to=date_to,
        section=section,
        branch=branch,
        min_percentage=min_percentage
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Summary"

    # --- Title Row ---
    title = "Student Attendance Summary"
    if date_from and date_to:
        title += f" — {date_from} to {date_to}"
    if min_percentage is not None:
        title += f" (Below {min_percentage}%)"
    ws.merge_cells('A1:G1')
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name='Calibri', bold=True, size=14, color='FF5722')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Headers ---
    headers = ['Student Name', 'USN (Roll No.)', 'Branch', 'Section',
               'Days Present', 'Total Days', 'Attendance %']
    _style_excel_header(ws, headers, row=3)

    # --- Data ---
    pct_fill_good = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    pct_fill_warn = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
    pct_fill_bad = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for idx, record in enumerate(data, start=4):
        ws.cell(row=idx, column=1, value=record['name'])
        ws.cell(row=idx, column=2, value=record['roll_number'])
        ws.cell(row=idx, column=3, value=record['branch'])
        ws.cell(row=idx, column=4, value=record['section'])
        ws.cell(row=idx, column=5, value=record['days_present'])
        ws.cell(row=idx, column=6, value=record['total_days'])
        pct_val = record['percentage']
        pct_cell = ws.cell(row=idx, column=7, value=f"{pct_val}%")

        # Style all cells in this row
        for col_idx in range(1, 8):
            cell = ws.cell(row=idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border

        # Color-code the percentage cell
        if pct_val >= 75:
            pct_cell.fill = pct_fill_good
            pct_cell.font = Font(name='Calibri', size=10, bold=True, color='2E7D32')
        elif pct_val >= 50:
            pct_cell.fill = pct_fill_warn
            pct_cell.font = Font(name='Calibri', size=10, bold=True, color='E65100')
        else:
            pct_cell.fill = pct_fill_bad
            pct_cell.font = Font(name='Calibri', size=10, bold=True, color='C62828')

    # --- Auto-fit column widths ---
    for col_idx in range(1, 8):
        max_len = 0
        for row_idx in range(3, 4 + len(data)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    # --- Send file ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "attendance_summary"
    if date_from and date_to:
        filename += f"_{date_from}_to_{date_to}"
    if min_percentage is not None:
        filename += f"_below_{int(min_percentage)}pct"
    filename += ".xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
